#!/usr/bin/python3
# -*-coding:Utf-8 -*

#     BIDS Pipeline select and analyse data in BIDS format.
#     Copyright © 2018-2020 Aix-Marseille University, INSERM, INS
#
#     This file is part of BIDS Pipeline. This file creates statistical table.
#
#     BIDS Pipeline is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     any later version
#
#     BIDS Pipeline is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with BIDS Pipeline.  If not, see <https://www.gnu.org/licenses/>
#
#     Authors: Aude Jegou, 2019-2021

import os
from scipy.io import loadmat
from numpy import ndarray
import json
import xlrd
import openpyxl
from datetime import datetime

# variables used in all script
# 'anat', 'pet', 'func', 'fmap', 'dwi',
__modality_type__ = ['ieeg', 'eeg', 'meg', 'beh']
__channel_name__ = ['channel', 'channels', 'electrode_name',
                    'electrodes_name', 'electrode_names', 'label', 'labels']
__file_attr__ = ['sub', 'ses', 'task', 'acq', 'proc', 'run']
__reject_dir__ = ['parsing', 'log', 'parsing_old',
                  'log_old', 'bids_pipeline', 'bids_uploader', 'anywave']


def go_throught_dir_to_convert(dirname, modality, subject_selected):
    """ Parse the directory and convert the files in tsv if it's possible

    :param dirname: str, path of the directory to parse
    :param modality: list, List of modality that can be parsed
    :param subject_selected: list, List of subject that can be parsed
    :return: Log with error and list of tsv files (to put on the table)
    """
    header = None
    if dirname.endswith('gardel'):
        modality = 'anat'

    if modality == 'anat':  # default values
        header = ["channel", "gardel_regions"]

    log_error = ''
    tmp_list = []
    if not os.path.exists(dirname):
        raise ('The directroy doesn"t exists')
    with os.scandir(dirname) as it:
        for entry in it:
            if ((entry.name.startswith('sub-') and entry.name in subject_selected) or entry.name.startswith('ses-')) and entry.is_dir():
                log_error, tsv_list = go_throught_dir_to_convert(
                    entry.path, modality, subject_selected)
                tmp_list.extend(tsv_list)
            elif entry.name in modality and entry.is_dir():
                file_list = os.listdir(entry.path)
                ext_list = [os.path.splitext(os.path.join(entry.path, file))[
                    1] for file in file_list]
                ext_list = list(set(ext_list))
                if '.tsv' in ext_list:
                    tmp_list = [os.path.join(
                        entry.path, file) for file in file_list if os.path.splitext(file)[1] == '.tsv']
                    return '', tmp_list
                elif '.mat' in ext_list and ('.xls' and '.xlsx') not in ext_list:
                    mat_file = [file for file in file_list if '.mat' in file]
                    tmp_list = []
                    for file in mat_file:
                        filename = os.path.join(entry.path, file)
                        try:
                            json_dict, tsv_dict = convert_mat_file(filename)
                            log_error, tsv_file = write_file_after_convert(
                                json_dict, tsv_dict)
                            if tsv_file is not None:
                                if isinstance(tsv_file, str):
                                    tmp_list.append(
                                        os.path.join(entry.path, tsv_file))
                                elif isinstance(tsv_file, list):
                                    for f in tsv_file:
                                        tmp_list.append(
                                            os.path.join(entry.path, f))
                        except:
                            log_error += filename + ' could not be converted in tsv'
                    return log_error, tmp_list
                elif '.xlsx' in ext_list or '.xls' in ext_list:
                    xls_file = [file for file in file_list if '.xls' in file]
                    tmp_list = []
                    for file in xls_file:
                        filename = os.path.join(entry.path, file)
                        try:
                            json_dict, tsv_dict = convert_xls_file(
                                filename, header)
                            log_error, tsv_file = write_file_after_convert(
                                json_dict, tsv_dict)
                            if tsv_file is not None:
                                if isinstance(tsv_file, str):
                                    tmp_list.append(
                                        os.path.join(entry.path, tsv_file))
                                elif isinstance(tsv_file, list):
                                    for f in tsv_file:
                                        tmp_list.append(
                                            os.path.join(entry.path, f))
                        except:
                            log_error += filename + ' could not be converted in tsv format'
                            continue
                    return log_error, tmp_list
                else:
                    log_error += 'Some files do not have the right extension (i.e tsv, mat, tsv) and cannot be converted.\n'
    return log_error, tmp_list


def parse_derivatives_folder(folder, table2write, stack=False, folder_in=None, derivatives_folder=None, modality=None, subject_selected=None):
    """Parse the derivatives folder and write the table

    :param folder: str, path
    :param table2write: dict, to create the table
    :param stack: boolean, determine the type of table (i.e. stack or unstack)
    :param folder_in: list, module results that should be present in the table
    :param derivatives_folder: str, path of the derivative folder (seems redundant with foler bt it's for the split,
    I need to know for sure the derivative path)
    :param modality: list, List of modality that can be parsed
    :param subject_selected: list, List of subject that can be parsed
    :return: Log with error
    """
    if subject_selected is None:
        log_error = 'There is no subject selected.\n'
        return log_error
    dev_folder = None
    if derivatives_folder:
        dev_folder = derivatives_folder
    log_error = ''
    if folder_in is None:
        folder_in = [elt for elt in os.listdir(
            dev_folder) if elt not in __reject_dir__]
    header = None
    with os.scandir(folder) as it:
        for entry in it:
            if entry.name in folder_in and entry.is_dir():
                # Find the software name:
                soft_name = entry.path.split(dev_folder)[1]
                soft_name = soft_name.split('\\')[1]
                error, file_list = go_throught_dir_to_convert(
                    entry.path, modality, subject_selected)
                log_error += error
                #file_list = os.listdir(entry.path)
                for file in file_list:
                    json_dict, tsv_dict = read_tsv(file)
                    key2remove = []
                    for key in tsv_dict:
                        if key.lower() in __channel_name__ and all(isinstance(elt, str) for elt in tsv_dict[key]):
                            channel = tsv_dict[key]
                            key2remove.append(key)
                    try:
                        header = create_table(
                            tsv_dict, channel, key2remove, table2write, header, stack, file, soft_analysis=soft_name)
                    except UnboundLocalError:
                        log_error += 'The file {} cannot be present in the table.\n'.format(
                            file)
                        pass
            # elif entry.name not in folder_out and entry.is_dir():
            #     error = parse_derivatives_folder(entry.path, table2write, table_attributes, derivatives_folder=dev_folder)
            #     log_error += error
    return log_error


def write_big_table(derivatives_folder, output_dir, folder_in=None, modality=None, subject_selected=None, stack=True):
    """Write the statistical table

    :param derivatives_folder: str, path of the derivatives folder
    :param folder_in: list, module results that should be present in the table
    :param modality: list, List of modality that can be parsed
    :param subject_selected: list, List of subject that can be parsed
    :param stack: boolean, determine the type of table (i.e. stack or unstack)
    :param output_dir: output directory, default should be derivatives
    :return: Log with error
    """

    if not output_dir:
        raise ValueError

    if modality is None:
        modality = __modality_type__
        stmod = 'all'
    elif isinstance(modality, str):
        stmod = modality
        modality = [modality]
    if subject_selected is None:
        log_error = 'No subjects have been selected.\n'
        return log_error
    if stack:
        stackname = 'stack'
    else:
        stackname = 'unstack'
    date = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
    # before write_table
    table2write = [[]]
    log_error = parse_derivatives_folder(derivatives_folder, table2write, stack=stack, folder_in=folder_in,
                                         derivatives_folder=derivatives_folder, modality=modality, subject_selected=subject_selected)
    filename = 'statistical_table_' + stmod + '_' + stackname + '_' + date + '.tsv'
    header_len = len(table2write[0])
    if header_len > 0:
        for line in table2write[1::]:
            if len(line) != header_len:
                diff_len = header_len-len(line)
                line.extend(['n/a']*diff_len)
        write_table(table2write, filename, output_dir)
        log_error += 'The statistical table has been written in derivatives folder with the following name: {}.\n'.format(
            filename)
    return log_error


def convert_mat_file(filename):
    """Read matlab file and convert it in tsv file

    :param filename: str, name of the file
    :return: dictionary with information, ictionary of the table
    """
    def turn_array_in_list(elt):
        """ Convert array (matrice) in list

        :param elt: array
        :return: List of the array
        """
        if isinstance(elt, ndarray):
            value = elt.tolist()
            final_value = turn_array_in_list(value)
        elif isinstance(elt, list):
            value = []
            for val in elt:
                if isinstance(val, list):
                    inter_value = turn_array_in_list(val)
                    value.extend(inter_value)
                elif isinstance(val, ndarray):
                    if val.dtype.fields is None:
                        if val.size == 1:
                            value.append(val.item())
                        elif val.size <= 306:
                            val = val.tolist()
                            inter_value = turn_array_in_list(val)
                            value.extend(inter_value)
                else:
                    value.append(val)
            final_value = value
        return final_value

    data = loadmat(filename)
    json_dict = {}
    tsv_dict = {}
    mat_keys = ['__header__', '__version__', '__globals__']
    keylist = [key for key in data.keys() if key not in mat_keys]
    for key in keylist:
        if data[key].dtype.hasobject and data[key].dtype.fields is not None:
            dt = data[key].dtype
            for stname in dt.names:
                value = turn_array_in_list(data[key][stname])
                if isinstance(value, str):
                    json_dict[stname] = value
                elif isinstance(value, list):
                    if len(value) == 1:
                        json_dict[stname] = value[0]
                    else:
                        tsv_dict[stname] = value
                else:
                    tsv_dict[stname] = value
        else:
            if data[key].size == 1:
                json_dict[key] = data[key].item()
            elif data[key].size > 1:
                shape = data[key].shape
                if len(shape) <= 2:
                    value = turn_array_in_list(data[key])
                    tsv_dict[key] = value
                else:
                    json_dict[key] = 'value is in {} dimension'.format(
                        len(shape))
    json_dict['filename'] = filename
    return json_dict, tsv_dict


def convert_xls_file(filename, header=None):
    """Read excel file and convert it in tsv file

    :param filename: str, name of the file
    :return: dictionary with information, ictionary of the table
    """
    def check_header_continuity(tmp_dict):
        """Check if the row is the header or not (without blank between columns)
        I think that checks header continuity (no empty columns) by seeing the index.

        :param tmp_dict: dict of the row
        :return:
        """
        nbr = [key for key in tmp_dict]
        if len(nbr) > 0:
            all_true = []
            last_elt = nbr[0]
            for elt in nbr[1:]:
                if elt == last_elt + 1:
                    all_true.append(True)
                else:
                    all_true.append(False)
                last_elt = elt
            if all(all_true):
                return True
            elif all_true.count(False) < round(len(nbr)*10/100):
                return True
            else:
                return False
        else:
            return False

    def check_header_values(tmp_dict, header=[]):
        """"Checks if the header contains one of the possible values (should be more, but for now is for gardel)."""
        POSSIBLE_HEADERS = ["Channels", "gardel_regions",
                            "channels", "brain_area"]+header
        #POSSIBLE_HEADERS=["Channels", "gardel_regions"]+header

        for key, value in tmp_dict.items():
            if value in POSSIBLE_HEADERS:
                return True

        return False

    # for excel's extension compatibility
    excel_xls = False
    if filename.endswith('xls'):
        excel_xls = True

    if excel_xls:
        workbook = xlrd.open_workbook(filename)
        nsheets = workbook.nsheets
    else:
        workbook = openpyxl.load_workbook(filename, data_only=True)
        nsheets = len(workbook.sheetnames)

    json_dict = {}
    tsv_dict = {}
    for i in range(0, nsheets, 1):
        json_dict[i] = {}
        if excel_xls:
            worksheet = workbook.sheet_by_index(i)
            ncol = worksheet.ncols
            nrow = worksheet.nrows
        else:
            worksheet = workbook.worksheets[i]
            ncol = worksheet.max_column
            nrow = worksheet.max_row

        if ncol == 0:
            continue

        if excel_xls:
            iterator = enumerate(worksheet._cell_values)
        else:
            iterator = enumerate(worksheet.rows)

        ignoring_header_index = 0
        for j, row in iterator:
            if excel_xls:
                tmp_dict = {worksheet._cell_values[j].index(
                    value): value for value in worksheet._cell_values[j] if value}
            else:
                tmp_dict = {
                    row.index(cell): cell.value for cell in row if cell}

            if check_header_continuity(tmp_dict):
                if header:
                    if check_header_values(tmp_dict, header):
                        if "gardel_regions" in header:
                            for key, value in tmp_dict.items():
                                if value == 'brain_area':
                                    # hardcore
                                    tmp_dict[key] = "gardel_regions"

                        ignoring_header_index = j + 1
                        break
                    else:
                        # set the default header
                        for idx, key in enumerate(tmp_dict.keys()):
                            tmp_dict[key] = header[idx]

                        break

                else:
                    ignoring_header_index = j + 1
                    break

        tsv_tmp_dict = {val: [] for key, val in tmp_dict.items()}
        all_blank = False
        datemode = 1
        if excel_xls:
            enumerator = enumerate(
                worksheet._cell_values[ignoring_header_index:])
        else:
            enumerator = enumerate(worksheet.rows)

        for nbr, row in enumerator:
            if not excel_xls:
                if nbr < ignoring_header_index:
                    continue

                nbr = nbr-ignoring_header_index

            if excel_xls:
                all_blank = all(elt == '' for elt in row)
            else:
                all_blank = all(elt.value == '' or elt.value ==
                                None for elt in row)

            if not all_blank:
                for cnt, elt in enumerate(row):
                    if not excel_xls:
                        elt = elt.value

                    if cnt < len(tmp_dict) and cnt in tmp_dict:
                        if tmp_dict[cnt] == 'DOB' and elt != '':
                            dob = datetime.strptime(elt, '%d\%m\%Y')
                            elt = dob.strftime('%d%m%Y')
                        elif isinstance(elt, float) and tmp_dict[cnt].lower().startswith('date'):
                            if not excel_xls:
                                if workbook.epoch == openpyxl.utils.datetime.CALENDAR_WINDOWS_1900:
                                    datemode = 0
                            else:
                                datemode = workbook.datemode

                            date = datetime(
                                *xlrd.xldate_as_tuple(elt, datemode))
                            elt = date.strftime('%d/%m/%Y')
                        tsv_tmp_dict[tmp_dict[cnt]].append(elt)
                    elif cnt not in tmp_dict and cnt-1 in tmp_dict:
                        if len(tsv_tmp_dict[tmp_dict[cnt-1]]) <= nbr:
                            tsv_tmp_dict[tmp_dict[cnt - 1]].append(elt)
                        elif not tsv_tmp_dict[tmp_dict[cnt - 1]][-1]:
                            tsv_tmp_dict[tmp_dict[cnt - 1]][-1] = elt
            else:
                # ignore if it is an empty row
                if excel_xls:
                    idx = worksheet._cell_values.index(row) + 1
                    iterator = worksheet._cell_values[idx::]
                else:
                    idx = nbr + 1
                    iterator = enumerate(worksheet.rows)

                for r_idx, r in iterator:
                    # we start from the next row after the empty one
                    if not excel_xls and r_idx <= idx:
                        continue

                    if excel_xls:
                        all_blank = all(elt == '' for elt in r)
                    else:
                        all_blank = all(
                            elt.value == '' or elt.value == None for elt in r)

                    if not all_blank:
                        col_indices = []
                        if excel_xls:

                            col_indices = [j for j, e in enumerate(
                                r) if e != '' and e != None]
                            count = len(col_indices)
                        else:
                            col_indices = [j for j, e in enumerate(
                                r) if e.value != '' and e.value != None]
                            count = len(col_indices)

                        if count % 2 == 0:
                            for j in range(0, len(col_indices), 2):
                                if excel_xls:
                                    if r[col_indices[j]] and j+1 < ncol:
                                        json_dict[i][r[col_indices[j]]
                                                     ] = r[col_indices[j+1]]
                                else:
                                    if r[j].value and j+1 < ncol:
                                        json_dict[i][r[col_indices[j]
                                                       ].value] = r[col_indices[j+1]].value
                break

        create_name = False
        try:
            if 'filename' not in json_dict[i]:
                create_name = True
            elif 'sub' not in json_dict[i]["filename"]:
                create_name = True
        except Exception as err:
            raise err

        if create_name:
            if excel_xls:
                suffix = worksheet.name
            else:
                suffix = worksheet.title

            suffix = suffix.replace(' ', '')
            suffix = suffix.replace('_', '')
            file_list = os.path.basename(filename).split('_')
            prefix = ''
            if 'sub' not in file_list[0]:
                prefix = 'sub-'

            # creating filename of the tsv
            json_dict[i]['filename'] = prefix + \
                '_'.join(file_list[:-2]) + '_' + suffix + '.tsv'

        # this is a little shit, but it solves when the excel sheet is not read properly (we don't know why)
        # and we have an emtpy column
        tsv_dict[i] = {}
        for k, v in tsv_tmp_dict.items():
            if k is not None and not all(elem == None for elem in v):
                tsv_dict[i][k] = v

    if len(tsv_dict.keys()) == 1:
        tsv_dict = tsv_dict[0]
        json_dict = json_dict[0]
    json_dict['filename'] = filename

    return json_dict, tsv_dict


def convert_txt_file(filename):
    """Read text file and convert it in tsv file

    :param filename: str, name of the file
    :return: dictionary with information, ictionary of the table
    """
    json_dict = {'filename': filename}
    f = open(filename, 'r+')
    f_cont = f.readlines()
    f_cont = [line.replace('\n', '') for line in f_cont]
    if all('\t' in line for line in f_cont):
        pass
    # to complicated to handle I guess
    # Recuperer les lignes avec separateur /n
    # si dans la ligne /t
    # en faire un tableau
    pass


def read_tsv(filename):
    """Read TSV file

    :param filename: str, name of the file
    :return: dictionary with information, dictionary of the table
    """
    json_dict = {'filename': filename}
    file = open(filename, 'r')
    data = file.readlines()
    data = [line.replace('\n', '') for line in data]
    if data:
        header = data[0].split('\t')
        tsv_dict = {key: [] for key in header}
        for lines in data[1::]:
            line = lines.split('\t')
            for key in header:
                idx = header.index(key)
                if idx < len(line):
                    tsv_dict[key].append(line[idx])
                else:
                    tsv_dict[key].append('')
    else:
        tsv_dict = {}

    return json_dict, tsv_dict


def read_vmrk(filename):
    """Read vmrk file

    :param filename: str, name of the file
    :return: dictionary with information, dictionary of the table
    """
    infos_dict = extract_info_vhdr(filename.replace('.vmrk', '.vhdr'))
    ts = float(infos_dict['SamplingInterval'])*1e-6
    json_dict = {'filename': filename}
    tsv_dict = {'onset': [], 'duration': [], 'trial_type': []}
    file = open(filename, 'r')
    data = file.readlines()
    if '[Marker Infos]\n' in data:
        beg = data.index('[Marker Infos]\n')
        for line in data[beg+1::]:
            if line.startswith('Mk'):
                elt = line.split(',')
                tsv_dict['trial_type'].append(elt[1])
                tsv_dict['onset'].append(str(float(elt[2])*ts))
                tsv_dict['duration'].append(str(float(elt[3])*ts))

    file.close()
    return json_dict, tsv_dict


def extract_info_vhdr(filename):
    """Read vhdr file and get the information

    :param filename: str, name of the file
    :return: dictionary with information
    """
    file = open(filename, 'r')
    data = file.readlines()
    beg = data.index('[Common Infos]\n')
    fin = data.index('[Binary Infos]\n')
    infos_dict = {}
    for line in data[beg+1:fin]:
        line = line.replace('\n', '')
        if '=' in line:
            elt = line.split('=')
            infos_dict[elt[0]] = elt[1]
    file.close()
    return infos_dict


def create_table(tsv_dict, channel, key2remove, table2write, header_real_val=None, stack=False, file=None, soft_analysis=None):
    """Create table according to the dictionary tsv_dict

    :param tsv_dict: dict, with the value to put in the table
    :param channel: list, channels name
    :param key2remove: list, elements to not add to the table
    :param table2write: dict, table resulting (push the value into it)
    :param header_real_val: dict, header with the original name and the new one
    :param stack: boolean, type of the table
    :param file: str, name of the file
    :param soft_analysis: str, name of the module
    :return: dict, header with the original name and the new one
    """
    if header_real_val is None:
        header_real_val = {'header': table2write[0]}
    name = {}
    if soft_analysis:
        soft = soft_analysis + '_'
    else:
        soft = ''

    if file is not None:
        # extracts new columns for the table
        name_list = os.path.basename(file).split('_')
        name = {elt.split('-')[0]: elt.split('-')[1] for elt in name_list if
                '-' in elt and elt.split('-')[0] in __file_attr__}
    # else:
    #     name = {}
    if stack:
        if not header_real_val['header']:
            header_real_val['header'].extend(name)
        elif any(elt not in header_real_val['header'] for elt in name):
            if 'Channel' in header_real_val['header']:
                idx = header_real_val['header'].index('Channel')
            else:
                idx = len(header_real_val['header'])
            for key in name.keys():
                if key not in header_real_val['header']:
                    # we add the new columns related with the file name
                    header_real_val['header'].insert(idx, key)
                    for line in table2write[1::]:
                        # this will provoke the stack behaviour
                        line.insert(idx, 'n/a')
    else:
        col_attr = [name[nm]
                    for nm in name if nm in __file_attr__ and not nm == 'sub']
        key_attr = '_'.join(col_attr)

    if 'sub' not in header_real_val['header'] and file is not None:
        header_real_val['header'].insert(0, 'sub')
    if 'Channel' not in header_real_val['header']:
        header_real_val['header'].insert(
            len(header_real_val['header']), 'Channel')
        # col_attr = [name[nm] for nm in name if nm in __file_attr__]
        # key_attr = '_'.join(col_attr)

    # save tsv columns to fill this ones later
    tsv_columns = list()
    for key in tsv_dict:
        if stack:
            key2write = key.replace(' ', '_')
        else:
            key2write = key_attr + '_' + key.replace(' ', '_')

        if (key not in key2remove and key != ''
                and len(tsv_dict[key]) == len(channel)):
            new_key = soft + key2write
            tsv_columns.append(new_key)
            if soft + key2write not in header_real_val['header']:
                header_real_val['header'].append(new_key)
                header_real_val[new_key] = {'infile': key, 'softname': soft}

    for i in range(0, len(channel), 1):
        chan = channel[i]
        id_chan = header_real_val['header'].index('Channel')
        flag_not_inside = True
        lines = None
        if table2write:
            # checks if the fields in 'name' and the channel of the row match
            # to update the rest of the fields, if not, it will be stacked
            idx = [cnt for cnt, l in enumerate(table2write) if all(l[j] == name[header_real_val['header'][j]] for j in range(
                0, id_chan, 1) if header_real_val['header'][j] in name) and l[id_chan] == chan]
            if idx:
                flag_not_inside = False
                lines = table2write[idx[0]]
                if len(lines) != len(header_real_val['header']):
                    for n in range(len(lines), len(header_real_val['header']), 1):
                        lines.append('n/a')
            else:
                lines = ['n/a'] * len(header_real_val['header'])
        else:
            lines = ['n/a'] * len(header_real_val['header'])

        for idx, key in enumerate(header_real_val['header']):
            try:
                if key in header_real_val and (header_real_val[key]['infile'] in tsv_dict and header_real_val[key]['softname'] == soft):
                    if key in tsv_columns:
                        lines[idx] = str(
                            tsv_dict[header_real_val[key]['infile']][i])
                elif key == 'Channel' and lines[idx] == 'n/a':
                    lines[idx] = chan
                elif key in name.keys() and lines[idx] == 'n/a':
                    lines[idx] = str(name[key])
                else:
                    continue
            except:
                continue
        if flag_not_inside:
            table2write.append(lines)
    # check the size of the lines compare to the header
    ncol = len(header_real_val['header'])
    if any(len(elt) != ncol for elt in table2write):
        for lines in table2write:
            ntoadd = ncol - len(lines)
            lines.extend(['n/a']*ntoadd)

    return header_real_val


def write_table(table2write, filename, output_dir, json_dict=None):
    """Write the table in TSV format and the JSON associated if exist

    :param table2write: dict, table to write
    :param filename: str, name of the file
    :param output_dir: str, output path
    :param json_dict: dict, dictionary with inforamtion
    :return:
    """
    if json_dict:
        if 'filename' in json_dict:
            file, ext = os.path.splitext(json_dict['filename'])
            jsonfilename = file + '.json'
        else:
            jsonfilename = os.path.splitext(filename)[0] + '.json'
        if os.path.exists(jsonfilename) and os.path.getsize(jsonfilename) != 0:
            with open(jsonfilename, 'r') as file:
                exist_dict = json.load(file)
                file.close()
        else:
            exist_dict = {key: '' for key in [
                'Description', 'RawSources', 'Parameters', 'Author', 'Date']}
            exist_dict['Date'] = datetime.now().strftime("%Y-%m-%dT%H-%M-%S")
        if output_dir not in jsonfilename:
            jsonfilename = os.path.join(output_dir, jsonfilename)
        with open(jsonfilename, 'w') as f:
            if 'Parameters' not in exist_dict.keys():
                exist_dict['Parameters'] = json_dict
            else:
                if isinstance(exist_dict['Parameters'], dict):
                    for clef, val in json_dict.items():
                        exist_dict['Parameters'][clef] = val
                else:
                    exist_dict['Parameters'] = json_dict
            json_str = json.dumps(exist_dict, indent=1, separators=(
                ',', ': '), ensure_ascii=False, sort_keys=False)
            f.write(json_str)
    # write tsv file
    if output_dir not in filename:
        tsvfilename = os.path.join(output_dir, filename)
    else:
        tsvfilename = filename
    with open(tsvfilename, 'w') as file:
        for lines in table2write:
            file.write('\t'.join(lines) + '\n')


def write_file_after_convert(json_dict, tsv_dict):
    """Write the resulting file TSV and JSON after conversion

    :param json_dict: dict, information about the file
    :param tsv_dict: dict, table
    :return: log error and the filename of file
    """
    log_error = ''
    flag_multifile = all(isinstance(key, int) for key in tsv_dict.keys())
    key2remove = []

    if flag_multifile:
        filelist = []
        for i in tsv_dict:
            sheet_name = ''
            for key in tsv_dict[i]:
                if key.lower() in __channel_name__ and all(isinstance(elt, str) for elt in tsv_dict[i][key]):
                    channel = tsv_dict[i][key]
                    key2remove.append(key)
            if not key2remove:
                continue
            try:
                table2write = [[]]
                if 'filename' in json_dict[i]:
                    filename = os.path.splitext(
                        json_dict[i]['filename'])[0] + '.tsv'
                    sheet_name = filename.split('_')[-1]
                    sheet_name = sheet_name.split('.')[0]
                else:
                    filename = os.path.splitext(
                        filename)[0] + '_run-' + str(i) + '.tsv'
                    sheet_name = filename.split('_run')[0]
                    sheet_name = sheet_name.split('_')[-1]

                path = os.path.dirname(filename)
                if not path:
                    path = os.path.dirname(json_dict['filename'])
                header = create_table(
                    tsv_dict[i], channel, key2remove, table2write, stack=True)
                if table2write[0] != ['Channel'] and len(table2write) > 1:
                    write_table(table2write, filename, path,
                                json_dict=json_dict[i])
                    filelist.append(filename)
                else:
                    log_error += 'The table {} for the file {} cannot be created.\n'.format(
                        sheet_name, json_dict['filename'])
                    filename = None
            except:
                log_error += 'The table for the file {} cannot be created.\n'.format(
                    json_dict['filename'])
                filename = None
        filename = filelist
    else:
        filename = os.path.splitext(json_dict['filename'])[0] + '.tsv'
        path = os.path.dirname(filename)
        table2write = [[]]
        for key in tsv_dict:
            if key.lower() in __channel_name__ and all(isinstance(elt, str) for elt in tsv_dict[key]):
                channel = tsv_dict[key]
                key2remove.append(key)
        try:
            header = create_table(
                tsv_dict, channel, key2remove, table2write, stack=True)
            if table2write[0] != ['Channel'] and len(table2write) > 1:
                write_table(table2write, filename, path, json_dict)
            else:
                log_error += 'The table for the file {} cannot be created.\n'.format(
                    json_dict['filename'])
                filename = None
        except:
            log_error += 'The table for the file {} cannot be created.\n'.format(
                json_dict['filename'])
            filename = None
    return log_error, filename


def convert_channels_in_montage_file(channelfile, modalitytype, outdirname=None):
    """Convert the channels file (channels.tsv) in mtg file (anywave format)

    :param channelfile: str, filename of channels.tsv
    :param modalitytype: str, modality type
    :param outdirname: str, path of the output directory
    :return: log with error
    """
    err = ''
    dirname, filename = os.path.split(channelfile)
    f = open(channelfile, 'r')
    f_cont = f.readlines()
    f.close()
    if modalitytype.lower() in ['ieeg', 'eeg']:
        newfilename = filename.replace(
            'channels.tsv', modalitytype.lower() + '.vhdr.mtg')
    elif modalitytype.lower() in ['meg']:
        megdirname = os.path.join(dirname, filename.replace(
            'channels.tsv', modalitytype.lower()))
        if os.path.exists(megdirname):
            dirname = megdirname
            newfilename = 'c,rfDC.mtg'
        else:
            err += 'Cannot create the montage file for {} modality.\n'.format(
                modalitytype)
            return err
    if outdirname is not None:
        dirname = outdirname
    mtgfile = os.path.join(dirname, newfilename)
    if not os.path.exists(mtgfile):
        with open(mtgfile, 'w') as file:
            file.write('<!DOCTYPE AnyWaveMontage>\n<Montage>\n')
            for line in f_cont[1::]:
                sline = line.split('\t')
                channame = sline[0]
                modtype = sline[1]
                file.write(
                    '\t<Channel name="{}">\n\t\t<type>{}</type>\n\t\t<reference></reference>\n\t\t<color>black</color>\n\t</Channel>\n'.format(channame, modtype))
            file.write('</Montage>')
    return err


def convert_events_in_marker_file(eventfile, modalitytype, outdirname=None):
    """Convert the events file (vmrk or events.tsv) in mrk file (anywave format)

    :param eventfile: str, filename of events file
    :param modalitytype: str, modality type
    :param outdirname: str, path of the output directory
    :return: log with error
    """
    err = ''
    dirname, filename = os.path.split(eventfile)
    if eventfile.endswith('.tsv'):
        json_dict, tsv_dict = read_tsv(eventfile)
    elif eventfile.endswith('.vmrk'):
        newfilename = filename.replace('.vmrk', '.vhdr.mrk')
        json_dict, tsv_dict = read_vmrk(eventfile)
    else:
        err = 'The file format of {} is not supported'.format(filename)
        return err
    if modalitytype.lower() in ['ieeg', 'eeg']:
        newfilename = filename.replace(
            'events.tsv', modalitytype.lower() + '.vhdr.mrk')
    elif modalitytype.lower() in ['meg']:
        megdirname = os.path.join(dirname, filename.replace(
            'events.tsv', modalitytype.lower()))
        if os.path.exists(megdirname):
            dirname = megdirname
            newfilename = 'c,rfDC.mrk'
        else:
            err += 'Cannot create the marker file for {} modality.\n'.format(
                modalitytype)
            return err
    if outdirname is not None:
        dirname = outdirname
    mrkfile = os.path.join(dirname, newfilename)
    if not os.path.exists(mrkfile):
        with open(mrkfile, 'w') as file:
            file.write('// AnyWave Marker File\n')
            for cnt, elt in enumerate(tsv_dict['trial_type']):
                file.write('{0}\t-1\t{1}\t{2}\n'.format(elt,
                           tsv_dict['onset'][cnt], tsv_dict['duration'][cnt]))
    return err
